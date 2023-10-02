import logging
import os
import traceback

import pandas as pd
import regex as re
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import sql_field_report.constants.field_report_schema as schema
import sql_field_report.constants.illegal_chars as illegal_chars

logger = logging.getLogger(__name__)


def parse_file(file: str) -> str:
    """
    If file is a filepath return just the file name, otherwise return the original value
    Args:
        file: the file/table name

    Returns: str - the parsed value
    """
    try:
        value = os.path.basename(file)
    except:
        value = file

    return value


def generate_excel_report(analysis: pd.DataFrame, file_path: str) -> str:
    """Generate an excel data report

    Args:
        analysis (pd.DataFrame): the analysis dataframe
        filepath (str): the filepath to the produced excel

    Returns:
        str: the filepath of the produced excel
    """

    logger.info("Generating Excel Report...")

    analysis = analysis.replace(illegal_chars.ILLEGAL_CHARACTERS_RE, "", regex=True)
    analysis[schema.TABLE_FILE] = analysis[schema.TABLE_FILE].apply(
        lambda x: parse_file(x)
    )

    choice_values = analysis[analysis[schema.CHOICES] != ""][
        [
            schema.TABLE_FILE,
            schema.FIELD,
            schema.CHOICES,
        ]
    ]

    analysis = analysis[
        [
            schema.TABLE_FILE,
            schema.FIELD,
            schema.COUNT,
            schema.POPULATED,
            schema.UNIQUE,
            schema.DATATYPE,
            schema.TOP_VALUES,
        ]
    ]

    try:
        odd_fill = PatternFill(
            start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
        )
        even_fill = PatternFill(
            start_color="B8CCE4", end_color="B8CCE4", fill_type="solid"
        )

        gold_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )

        thin_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )

        with pd.ExcelWriter(file_path, engine="openpyxl") as xlsx:
            analysis.to_excel(xlsx, "Field Report", index=False)

            # Generate main report
            ws = xlsx.sheets["Field Report"]

            field_report_table = Table(
                displayName="FieldReport",
                ref="A1:{col}{row}".format(
                    col=get_column_letter(ws.max_column), row=ws.max_row
                ),
            )

            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )

            field_report_table.tableStyleInfo = style

            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 20
            ws.column_dimensions["G"].width = 100

            even = False
            for i, row in enumerate(ws.iter_rows()):
                ws["A{}".format(i + 1)].alignment = Alignment(vertical="top")
                ws["B{}".format(i + 1)].alignment = Alignment(vertical="top")
                ws["C{}".format(i + 1)].alignment = Alignment(vertical="top")
                ws["D{}".format(i + 1)].alignment = Alignment(vertical="top")
                ws["E{}".format(i + 1)].alignment = Alignment(vertical="top")
                ws["F{}".format(i + 1)].alignment = Alignment(vertical="top")
                ws["G{}".format(i + 1)].alignment = Alignment(wrapText=True)

                if i + 1 == 1:
                    pass
                else:
                    if ws["A{}".format(i + 1)].value != ws["A{}".format(i)].value:
                        even = not even
                    for cell in row:
                        cell.fill = even_fill if even else odd_fill

            ws.add_table(field_report_table)

            # Generate mapping tables
            # get tables with choice values
            tables_with_choices = (
                choice_values[schema.TABLE_FILE].drop_duplicates(keep="first").to_list()
            )

            for choice_table in tables_with_choices:
                sheet_name = re.sub(
                    illegal_chars.INVALID_TITLE_REGEX, "", f"Mappings- {choice_table}"
                )[:25]

                # get choice fields from table

                choices = choice_values[
                    choice_values[schema.TABLE_FILE] == choice_table
                ]
                choice_fields = (
                    choices[schema.FIELD].drop_duplicates(keep="first").to_list()
                )
                mapping_table_count = 0
                for choice_field in choice_fields:
                    mapping_table_data = choices[choices[schema.FIELD] == choice_field][
                        schema.CHOICES
                    ].to_list()[0]
                    mapping_table_data = list([[m, ""] for m in mapping_table_data])
                    mapping_table = pd.DataFrame.from_records(
                        data=mapping_table_data,
                        columns=[
                            f"{choice_field}- {schema.LEGACY}",
                            f"{choice_field}- {schema.TARGET}",
                        ],
                    ).replace(illegal_chars.ILLEGAL_CHARACTERS_RE, "", regex=True)
                    mapping_table = mapping_table[
                        mapping_table[f"{choice_field}- {schema.LEGACY}"] != ""
                    ]
                    mapping_table.to_excel(
                        xlsx,
                        sheet_name=sheet_name,
                        startcol=mapping_table_count
                        * schema.MAPPING_TABLE_COLUMNS_COUNT,
                        index=False,
                    )
                    choice_sheet = xlsx.sheets[sheet_name]
                    start_col = schema.calculate_col_index_for_mapping_table(
                        mapping_table_count, schema.LEGACY_COL_IDX
                    )
                    end_col = schema.calculate_col_index_for_mapping_table(
                        mapping_table_count, schema.TARGET_COL_IDX
                    )
                    options_col = schema.calculate_col_index_for_mapping_table(
                        mapping_table_count, schema.OPTIONS_COL_IDX
                    )

                    choice_sheet[
                        f"{get_column_letter(options_col)}1"
                    ].value = f"{choice_field}- Options"
                    choice_sheet[f"{get_column_letter(options_col)}1"].font = Font(
                        bold=True
                    )
                    choice_sheet[
                        f"{get_column_letter(options_col)}1"
                    ].border = thin_border
                    choice_sheet[
                        f"{get_column_letter(options_col)}2"
                    ].value = f"{{Insert CRM Options Here}}"

                    choice_sheet.column_dimensions[
                        get_column_letter(start_col)
                    ].width = "30"

                    choice_sheet.column_dimensions[
                        get_column_letter(end_col)
                    ].width = "30"

                    choice_sheet.column_dimensions[
                        get_column_letter(options_col)
                    ].width = "30"

                    for row in range(mapping_table.shape[0] + 1):
                        choice_sheet[
                            f"{get_column_letter(start_col)}{row+1}"
                        ].fill = even_fill
                        choice_sheet[
                            f"{get_column_letter(end_col)}{row+1}"
                        ].fill = even_fill
                        choice_sheet[
                            f"{get_column_letter(options_col)}{row+1}"
                        ].fill = gold_fill

                    mapping_table_count += 1

        logger.info("Excel Report Generated")

        return file_path

    except Exception as e:
        traceback.print_exc()
        logger.error(f"Excel Report Generation Failed: {e}")
        return ""
