import os

import pandas as pd
import regex as re
import typer
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import URL, create_engine, text


class MSSQLConnection(object):
    """Provides an MSSQL Connection

    Required Environment Variables:
        SQL_SERVER_PASSWORD (str): This is the password for connecting the the MS SQL Server

    Parameters:
        server (str): The server name/address
        port (int): The server port
        user (str): The username
        password (str): The passowrd
        db_name (str): The name of the database
    """

    def __init__(self, server: str, port: int, user: str, password: str, db_name: str):
        self._server = server
        self._port = port
        self._user = user
        self._password = password
        self._db_name = db_name
        self.engine = None
        self.conn = None

    def __enter__(self):
        connection_string = "DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server},{port};DATABASE={db};UID={uid};PWD={pwd};Trusted_Connection=No;".format(
            server=self._server,
            port=self._port,
            uid=self._user,
            pwd=self._password,
            db=self._db_name,
        )

        # Build connection URL
        connection_url = URL.create(
            "mssql+pyodbc", query={"odbc_connect": connection_string}
        )

        # SQL Server Engine
        self.engine = create_engine(connection_url, fast_executemany=True)

        self.conn = self.engine.connect()

        return self.conn

    def __exit__(self, type, value, traceback):
        self.conn.close()
        self.engine.dispose()


def most_common(lst):
    """
    Return the most common element in a list
    """
    return max(set(lst), key=lst.count)


def estimate_dealcloud_datatype(value, choice_flag):
    """
    Estimate the DealCloud datatype of a particular value

    Params:
    value
    bool: choice_flag

    Returns:
    str: datatype
    """
    value = str(value)
    dc_number_regex = r"^[0-9.,%xXmMbnB$£€GBPUSDEUR]+$"
    email_regex = r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)"
    date_regex = r"(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d{3}|\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}|\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}|\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}|\d{2,4}-\d{2}-\d{2,4}|\d{4}\/\d{2}\/\d{2} \d{2}:\d{2}:\d{2}\.\d{3}|\d{4}\/\d{2}\/\d{2} \d{2}:\d{2}:\d{2}|\d{2,4}\/\d{2}\/\d{2,4})"
    url_regex = r"[-a-zA-Z0-9@:%._\+~#=]{1,256}\.[a-zA-Z0-9()]{1,6}\b([-a-zA-Z0-9()@:%_\+.~#?&//=]*)"

    if choice_flag:
        return "Choice/Reference"
    elif re.search(dc_number_regex, value):
        # test if it is a DealCloud number field
        currency_indicators = ["m", "M", "bn", "B", "£", "$", "€", "GBP", "USD", "EUR"]
        if bool(
            [indicator for indicator in currency_indicators if (indicator in value)]
        ):
            return "Currency"
        elif "%" in value:
            return "Percentage"
        elif "x" in value.lower():
            return "Multiplier"
        else:
            return "Number"
    else:
        if value == "" or value == " ":
            return "EMPTY"
        elif re.search(email_regex, value):
            return "E-mail"
        elif re.search(date_regex, value):
            return "Date and Time"
        elif len(value) < 125:
            return "Single Line"
        elif len(value) > 125:
            return "Multi Line"
        else:
            return None


def get_choice_flag(distinct_count, choice_ratio, count) -> bool:
    """
    Determine if a field should be a choice or not

    Params:
    int: distinct_count - the unique values
    float: choice_ratio - the ratio of unique values:values
    int: count - the count of values

    Returns:
    bool: True if choice else false
    """
    if (
        (float(distinct_count) < 10 or float(choice_ratio) < 0.05)
        and (not float(distinct_count) == 0)
        and (not float(count) == float(distinct_count))
    ):
        return True
    else:
        return False


def review_column(file: str, data: pd.DataFrame, column: str):
    """
    Review a column of data in a data set to return field report metadata

    Params:
    str: file - The source file/table
    pd.DataFrame: data - the data
    str: column - the column being analyzed

    Return:
    tuple: analysis - the column analysis
    """

    if data.shape[0] != 0:
        length = len(data)
        values = data[column].dropna().to_list()
        populated = len(values)
        unique = len(set(values))

        if length > 0:
            choice_ratio = float(unique) / float(length)
        else:
            choice_ratio = length

        choice_flag = get_choice_flag(unique, choice_ratio, length)

        values = data[column].to_list()

        types = list([estimate_dealcloud_datatype(v, choice_flag) for v in set(values)])

        datatype = most_common(types)

        choices = (
            "; ".join(set(map(str, values))) if datatype == "Choice/Reference" else None
        )

        analysis = (file, column, length, populated, unique, datatype)
        # analysis = (file, column, length, populated, unique, datatype, choices)

        print("Analyzed: {}".format(str(analysis)))

    else:
        # analysis = (file, column, 0, 0, 0, "EMPTY", None)
        analysis = (file, column, 0, 0, 0, "EMPTY")

    return analysis


def analyze_file(table: str, schema: str, conn):
    """
    Take in a file path and return an overview of the shape of that file

    Params:
    str: table - the database table to query

    Returns:
    Tuple: file_shape - a tuple of tuples containing the file name, field name, row count for each field
    """
    try:
        data = pd.read_sql(text(f"SELECT * FROM [{schema}].[{table}]"), conn)

        file_shape = tuple((review_column(table, data, h) for h in data.columns))
    except:
        data = pd.DataFrame.from_records(data=[['ERROR', 'ERROR']], columns = ['ERROR', 'ERROR2'])
        file_shape = tuple((review_column(table, data, h) for h in data.columns))

    return file_shape


def analyze_files(objects: list, schema:str, conn) -> pd.DataFrame:
    """
    Analyze Files

    Params:
    list db_tables - list of database tables
    conn - sql server connection

    Returns:
    pd.DataFrame: analysis - a summary of all files, fields and their row counts
    """

    data_shapes = tuple(analyze_file(l, schema, conn) for l in objects)

    # flatten tuple
    data_shapes = tuple((element for t in data_shapes for element in t))

    analysis = pd.DataFrame(
        data=data_shapes,
        columns=[
            "Table/File",
            "Field",
            "Count",
            "Populated",
            "Unique",
            "Datatype",
            # "Choices",
        ],
    )

    return analysis


def build_field_report(output_file_name: str, objects: list, schema:str, conn):
    """
    Build field report

    Params:
    str: output_file_name - the name of the outputted file
    list: objects - list of tables to analyse
    conn: SQLAlchemy connection

    Returns:
    str: field_report - the field report file name

    Steps of Operation:
    If Database
    1. Produce tables and fields summary
        a. For DB execute a SQL query to return this info
    2. For each table:
        a. For each column, summarise:
            i. Number of blanks
            ii. Estimated data type
            iii. Possible choice values
    If Files
    1. Produce tables and fields summary
        a. Iterate through files and produce analysis on, for each column:
            i. File name
            ii. Column name
            iii. Row count
            iv. Number of blanks
            v. Estimated data type
            vi. Possible choice values
    """

    file_path = os.path.join(output_file_name)

    analysis = analyze_files(objects, schema, conn)
    # analysis.to_excel(file_path)

    oddFill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    evenFill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

    with pd.ExcelWriter(file_path, engine="openpyxl") as xlsx:
        analysis.to_excel(xlsx, "Field_Report")

        ws = xlsx.sheets["Field_Report"]

        ws.delete_cols(1, 1)

        field_report_table = Table(
            displayName="FieldReport",
            ref="A1:{col}{row}".format(
                col=get_column_letter(ws.max_column), row=ws.max_row
            ),
        )

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )

        field_report_table.tableStyleInfo = style

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 100

        even = False
        for i, row in enumerate(ws.iter_rows()):
            ws["A{}".format(i + 1)].alignment = Alignment(vertical="top")
            ws["B{}".format(i + 1)].alignment = Alignment(vertical="top")
            ws["C{}".format(i + 1)].alignment = Alignment(vertical="top")
            ws["D{}".format(i + 1)].alignment = Alignment(vertical="top")
            ws["E{}".format(i + 1)].alignment = Alignment(vertical="top")
            ws["F{}".format(i + 1)].alignment = Alignment(vertical="top")
            ws["G{}".format(i + 1)].alignment = Alignment(wrapText=True)

            if i + 1 == 1:
                pass
            else:
                print(
                    "{}: {} -> {}".format(
                        ws["B{}".format(i + 1)].value,
                        ws["A{}".format(i + 1)].value,
                        ws["A{}".format(i)].value,
                    )
                )
                if ws["A{}".format(i + 1)].value != ws["A{}".format(i)].value:
                    even = not even
                for cell in row:
                    cell.fill = evenFill if even else oddFill

        ws.add_table(field_report_table)

    return os.path.join("app", file_path)


app = typer.Typer()


@app.command()
def MSSQL_Database_Report(
    server: str,
    port: int,
    user: str,
    password: str,
    database_name: str,
    schema:str, 
    output_file_name: str,
):
    """MSSQL Database Report

    Generate an excel report summarising the data in an MSSQL Database

    Args:
        server (str): The SQL server name
        port (int): The SQL server port
        user (str): Your SQL Server username
        password (str): Your SQL Server password
        database_name (str): The name of the database to analyse
        schema (str): The schema for the database tables to analyse
        output_file_name (str): The output file name of the report
    """

    if not output_file_name.endswith(".xlsx"):
        output_file_name = "{}.xlsx".format(output_file_name.split(".")[0])

    with MSSQLConnection(server, port, user, password, database_name) as conn:
        objects = pd.read_sql(
            text("SELECT DISTINCT(TABLE_NAME) FROM INFORMATION_SCHEMA.COLUMNS"), conn
        )["TABLE_NAME"].to_list()
        build_field_report(output_file_name, objects, schema, conn)


if __name__ == "__main__":
    app()
